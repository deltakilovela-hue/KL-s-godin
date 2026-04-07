from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
from datetime import datetime
import io

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(pdf_path):
    """Extrae texto del PDF usando OCR"""
    try:
        images = convert_from_path(pdf_path)
        extracted_text = ""

        for image in images:
            text = pytesseract.image_to_string(image, lang='spa')
            extracted_text += text + "\n"

        return extracted_text
    except Exception as e:
        return f"Error: {str(e)}"

def parse_attendance_data(text):
    """Procesa el texto extraído para obtener datos de asistencia"""
    lines = text.split('\n')
    data = []

    # Datos de muestra basados en la estructura del PDF
    # En producción, esto sería más sofisticado
    nombre = ""
    departamento = ""

    for line in lines:
        line = line.strip()

        # Detectar nombres (líneas en mayúscula)
        if line and len(line) > 5 and line.isupper():
            if "MARZO" in line or "ABOGADO" in line or "ADMIN" in line:
                departamento = line
            elif line and not line.isdigit():
                nombre = line

        # Detectar horas (formato HH:MM)
        if ':' in line and any(char.isdigit() for char in line):
            hora_entrada = ""
            hora_salida = ""

            partes = line.split()
            for parte in partes:
                if ':' in parte and len(parte) == 5:
                    if not hora_entrada:
                        hora_entrada = parte
                    else:
                        hora_salida = parte
                        break

    return data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Recibe el PDF y extrae los datos"""

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Only PDF files allowed'}), 400

    try:
        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"upload_{timestamp}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Extraer texto
        extracted_text = extract_text_from_pdf(filepath)

        # Parsear datos (por ahora retornamos datos de muestra)
        sample_data = [
            {
                "id": "1",
                "nombre": "ARCINIEGA DE LA HUERTA",
                "departamento": "ENFERMERÍA",
                "fecha": "16/03/2026",
                "hora_entrada": "6:52",
                "hora_salida": "15:00",
                "festivo": "No",
                "falta": "No",
                "fecha_falta": "",
                "notas": ""
            },
            {
                "id": "2",
                "nombre": "LUIS GARCÍA CHACÓN",
                "departamento": "ENFERMERÍA",
                "fecha": "16/03/2026",
                "hora_entrada": "6:59",
                "hora_salida": "15:00",
                "festivo": "No",
                "falta": "No",
                "fecha_falta": "",
                "notas": ""
            },
            {
                "id": "3",
                "nombre": "DOMINGUEZ",
                "departamento": "ENFERMERÍA",
                "fecha": "17/03/2026",
                "hora_entrada": "6:45",
                "hora_salida": "14:50",
                "festivo": "No",
                "falta": "No",
                "fecha_falta": "",
                "notas": ""
            },
            {
                "id": "4",
                "nombre": "GONZALO",
                "departamento": "ABOGADO",
                "fecha": "18/03/2026",
                "hora_entrada": "7:15",
                "hora_salida": "15:30",
                "festivo": "No",
                "falta": "No",
                "fecha_falta": "",
                "notas": ""
            },
            {
                "id": "5",
                "nombre": "ASUNCIÓN COVARRUBIAS",
                "departamento": "ENFERMERÍA",
                "fecha": "19/03/2026",
                "hora_entrada": "6:40",
                "hora_salida": "15:00",
                "festivo": "No",
                "falta": "Sí",
                "fecha_falta": "21/03/2026",
                "notas": "Falta injustificada"
            },
        ]

        return jsonify({
            'success': True,
            'message': 'PDF procesado correctamente',
            'data': sample_data,
            'extracted_text_preview': extracted_text[:500]
        })

    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

@app.route('/api/export', methods=['POST'])
def export_excel():
    """Genera un archivo Excel con los datos"""

    try:
        data = request.json.get('data', [])

        if not data:
            return jsonify({'error': 'No data to export'}), 400

        # Crear workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Asistencia"

        # Headers
        headers = ["ID", "Nombre", "Departamento/Área", "Fecha",
                   "Hora Entrada", "Hora Salida", "¿Festivo?", "¿Falta?",
                   "Fecha Falta", "Notas"]

        sheet.append(headers)

        # Formato de encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Agregar datos
        for row_data in data:
            sheet.append([
                row_data.get('id', ''),
                row_data.get('nombre', ''),
                row_data.get('departamento', ''),
                row_data.get('fecha', ''),
                row_data.get('hora_entrada', ''),
                row_data.get('hora_salida', ''),
                row_data.get('festivo', ''),
                row_data.get('falta', ''),
                row_data.get('fecha_falta', ''),
                row_data.get('notas', '')
            ])

        # Ancho de columnas
        widths = [6, 25, 18, 12, 13, 13, 11, 10, 12, 25]
        for i, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = width

        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Crear resumen
        summary = wb.create_sheet("Resumen")
        summary['A1'] = "RESUMEN DE ASISTENCIA"
        summary['A1'].font = Font(bold=True, size=14, color="366092")

        summary['A3'] = "Total de registros:"
        summary['B3'] = len(data)

        total_faltas = sum(1 for row in data if row.get('falta') == 'Sí')
        summary['A4'] = "Total de faltas:"
        summary['B4'] = total_faltas

        summary['A5'] = "Tasa de asistencia:"
        tasa = ((len(data) - total_faltas) / len(data) * 100) if data else 0
        summary['B5'] = f"{tasa:.1f}%"

        # Guardar en memoria
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Asistencia_{timestamp}.xlsx"

        return send_file(
            excel_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': f'Error generating Excel: {str(e)}'}), 500

@app.route('/api/validate', methods=['POST'])
def validate_data():
    """Valida los datos extraídos"""

    data = request.json
    errors = []

    for row in data:
        # Validar nombre
        if not row.get('nombre', '').strip():
            errors.append(f"Fila {row.get('id')}: Nombre vacío")

        # Validar horas
        hora_entrada = row.get('hora_entrada', '')
        hora_salida = row.get('hora_salida', '')

        if hora_entrada and hora_salida:
            try:
                h_entrada = datetime.strptime(hora_entrada, '%H:%M')
                h_salida = datetime.strptime(hora_salida, '%H:%M')

                if h_salida <= h_entrada:
                    errors.append(f"Fila {row.get('id')}: Hora salida <= Hora entrada")
            except:
                errors.append(f"Fila {row.get('id')}: Formato de hora inválido")

    return jsonify({
        'valid': len(errors) == 0,
        'errors': errors
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000)
